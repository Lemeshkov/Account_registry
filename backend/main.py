# backend/main.py
from fastapi import (
    FastAPI,
    File,
    UploadFile,
    HTTPException,
    Depends,
    BackgroundTasks,
    Form,
    Path
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from pathlib import Path
import uuid
import shutil
import logging
from collections import defaultdict
from pydantic import BaseModel
from typing import List, Dict, Any
from datetime import datetime 

# Исправляем импорты - убираем точки
from database import get_db, engine, SessionLocal
from models import Base, PaymentRegistry, InvoiceLine, DefectSheet, DefectSheetItem 
from parsers.excel_parser import ExcelParser
from parsers.invoice_parser import parse_invoice_from_pdf
from crud import (
    build_registry_from_batch,
    apply_invoice_ocr_to_registry,
    create_imported_request,
    create_payment_registry_item,
    create_history,
    update_registry_position,
    reorder_registry_batch,
)
from redis_manager import redis_manager
from fastapi import WebSocket, WebSocketDisconnect
import asyncio
from services.invoice_matcher import try_match_invoice
from services.invoice_buffer import (
    list_invoices,
    add_invoice,
    save_invoice_lines,
    mark_invoice_line_used,
    get_invoice,
)
# Новые импорты для дефектной ведомости
from parsers.defect_parser import parse_defect_sheet
from services.metal_calculator import metal_calculator, ProfileType
from crud import (
    create_defect_sheet,
    create_defect_sheet_items,
    get_defect_sheet,
    get_defect_sheet_by_batch,
    get_defect_sheet_items,
    update_defect_sheet_status,
    update_defect_sheet_item_calculation,
    mark_items_for_calculation,
    delete_defect_sheet,
    create_defect_sheet,
)
from sqlalchemy.orm import Session
from database import get_db
from typing import List, Optional
import uuid
from pathlib import Path
import shutil
from sqlalchemy import func  
from sqlalchemy.orm import Session
import asyncio
from models import DefectSheet, DefectSheetItem  

log = logging.getLogger(__name__)
# WebSocket Manager импорт
from websocket_manager import websocket_manager

# -------------------------------------------------------------------
# APP
# -------------------------------------------------------------------

app = FastAPI(title="Registry Control API", version="1.0.0")

# Pydantic модели для запросов
class ApplyInvoiceLineRequest(BaseModel):
    invoice_id: str
    line_no: int
    registry_id: int

class ApplyInvoiceMetadataRequest(BaseModel):
    invoice_id: str
    registry_id: int
    apply_fields: List[str] = ["contractor", "invoice_number", "invoice_date"]

class ManualInvoiceMatchRequest(BaseModel):
    batch_id: str
    registry_id: int
    invoice_id: str
    apply_type: str = "full"  # "full", "metadata_only", "amount_only"

class ReorderRequest(BaseModel):
    batch_id: str
    items: List[Dict[str, Any]]  # [{id: 1, position: 0}, {id: 2, position: 1}, ...]

class ApplyMultipleLinesRequest(BaseModel):
    invoice_id: str
    line_nos: List[int]
    registry_id: int
    batch_id: str

class ApplyAllLinesRequest(BaseModel):
    invoice_id: str
    registry_id: int
    batch_id: str    


# Модели для дефектной ведомости
class DefectSheetResponse(BaseModel):
    id: int
    batch_id: str
    file_name: str
    upload_date: datetime
    status: str
    total_items: int = 0
    
    model_config = {
        "from_attributes": True
    }

class DefectSheetItemResponse(BaseModel):
    id: int
    position: Optional[int]
    address: Optional[str]
    material_name: Optional[str]
    requested_quantity: Optional[float]
    weight_tons: Optional[float]
    profile_type: Optional[str]
    profile_params: Optional[Dict]
    calculated_meters: Optional[float]
    formula_used: Optional[str]
    is_calculated: bool
    selected_for_calculation: bool
    
    model_config = {
        "from_attributes": True
    }

class DefectSheetPreviewResponse(BaseModel):
    sheet_id: int
    batch_id: str
    file_name: str
    upload_date: datetime
    status: str
    period_start: Optional[datetime]
    period_end: Optional[datetime]
    total_items: int
    items: List[DefectSheetItemResponse]

class CalculationRequest(BaseModel):
    sheet_id: int
    item_ids: Optional[List[int]] = None  # если None - пересчитать все непосчитанные
    profile_type: str = "pipe"  # по умолчанию труба
    profile_params: Dict[str, Any] = {}

class CalculationResponse(BaseModel):
    sheet_id: int
    total_items: int
    calculated_items: int
    results: List[Dict[str, Any]]

class ExportRequest(BaseModel):
    sheet_id: int
    format: str = "excel"



@app.on_event("startup")
async def startup_event():
    """Действия при запуске приложения"""
    await redis_manager.connect()
    print("🚀 Application started with Redis support")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_UPLOAD_DIR = Path("uploads")
BASE_UPLOAD_DIR.mkdir(exist_ok=True)

# Директория для дефектных ведомостей
DEFECT_UPLOAD_DIR = Path("uploads/defect_sheets")
DEFECT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app.mount("/uploads", StaticFiles(directory=BASE_UPLOAD_DIR), name="uploads")

# -------------------------------------------------------------------
# BACKGROUND OCR TASK
# -------------------------------------------------------------------

def process_invoice_pdf_background(file_path: Path, batch_id: str):
    db = SessionLocal()
    try:
        # 1. Парсим PDF
        parsed = parse_invoice_from_pdf(file_path)
        if not parsed:
            log.error("[OCR] Failed to parse PDF: %s", file_path.name)
            return
        
        print(f"\n=== DEBUG: parse_invoice_from_pdf result ===")
        print(f"Contractor: {parsed.get('data', {}).get('contractor')}")
        print(f"Invoice number: {parsed.get('data', {}).get('invoice_number')}")
        print(f"Invoice date: {parsed.get('data', {}).get('invoice_date')}")
        
        # 2. Добавляем ID и информацию о batch
        invoice_id = str(uuid.uuid4())
        parsed["id"] = invoice_id
        parsed["batch_id"] = batch_id
        parsed["file"] = file_path.name
        
        # 3. ВСЕГДА сохраняем строки счета в базу
        if parsed.get("lines"):
            log.info("[OCR] Saving %d invoice lines to database", len(parsed["lines"]))
            save_invoice_lines(
                db,
                invoice_id=invoice_id,
                batch_id=batch_id,
                lines=parsed.get("lines", []),
            )
        else:
            log.warning("[OCR] No product lines found in invoice")
        
        # 4. ВСЕГДА сохраняем счет в буфер (для предпросмотра)
        print(f"\n=== DEBUG: Adding to buffer ===")
        print(f"Invoice ID: {invoice_id}")
        
        # ОБЕСПЕЧИВАЕМ что data есть
        if 'data' not in parsed:
            parsed['data'] = {}
        
        # Добавляем обязательные поля если их нет
        if 'invoice_full_text' not in parsed['data']:
            invoice_number = parsed['data'].get('invoice_number')
            invoice_date = parsed['data'].get('invoice_date')
            if invoice_number and invoice_date:
                parsed['data']['invoice_full_text'] = f"Счет на оплату № {invoice_number} от {invoice_date}"
            elif invoice_number:
                parsed['data']['invoice_full_text'] = f"Счет на оплату № {invoice_number}"
            elif invoice_date:
                parsed['data']['invoice_full_text'] = f"Счет от {invoice_date}"
        
        # Сохраняем в буфер
        add_invoice(parsed)
        
        # 5. Фиксируем только строки товаров, НЕ изменяем реестр
        db.commit()
        log.info("[OCR] Processing complete for %s", file_path.name)
        print(f"[DEBUG] Invoice saved to buffer: {invoice_id}")
        print(f"[DEBUG] NO automatic application to registry!")
        
        # Отправляем уведомление о завершении обработки
        try:
            websocket_manager.broadcast_to_batch(batch_id, {
                "type": "invoice_processed",
                "batch_id": batch_id,
                "invoice_id": invoice_id,
                "filename": file_path.name,
                "contractor": parsed.get('data', {}).get('contractor'),
                "status": "completed"
            })
        except Exception as e:
            log.error(f"Failed to send WebSocket notification: {e}")
        
    except Exception as e:
        db.rollback()
        log.exception("[OCR] Failed to process invoice %s: %s", file_path.name, str(e))
        print(f"\n=== DEBUG: EXCEPTION ===")
        print(f"Error processing invoice: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()
        print(f"[DEBUG] Database connection closed")

# -------------------------------------------------------------------
# APPLY INVOICE LINE 
# -------------------------------------------------------------------

@app.post("/invoice/apply-line")
def apply_invoice_line(payload: ApplyInvoiceLineRequest, db: Session = Depends(get_db)):
    log.info(
        "[APPLY_LINE] invoice_id=%s line_no=%s registry_id=%s",
        payload.invoice_id,
        payload.line_no,
        payload.registry_id,
    )

    # Проверяем существует ли строка счета
    line = (
        db.query(InvoiceLine)
        .filter(
            InvoiceLine.invoice_id == payload.invoice_id,
            InvoiceLine.line_no == payload.line_no,
            InvoiceLine.used.is_(False),
        )
        .first()
    )
    
    # Строка счета может быть уже использована или не существовать
    # Это не должно блокировать применение, просто логируем
    if not line:
        log.warning("[APPLY_LINE] invoice line not found or already used: invoice_id=%s, line_no=%s",
                   payload.invoice_id, payload.line_no)

    registry = db.query(PaymentRegistry).get(payload.registry_id)
    if not registry:
        log.warning("[APPLY_LINE] registry not found")
        raise HTTPException(404, "Registry item not found")

    # Получаем данные счета из буфера
    invoice = get_invoice(payload.invoice_id)
    if not invoice:
        log.error("[APPLY_LINE] invoice not found in buffer")
        raise HTTPException(404, "Invoice not found")

    log.info("[APPLY_LINE] OCR DATA FROM BUFFER: %s", invoice["data"])

    invoice_data = {
        "id": payload.invoice_id,
        "data": dict(invoice["data"]),
        "lines": invoice.get("lines", []),
        "confidence": invoice.get("confidence", 0)
    }

    # Применяем счет к строке реестра с указанием номера строки
    apply_invoice_ocr_to_registry(
        db, 
        registry.id, 
        invoice_data,
        apply_full_metadata=True,  # Применяем все метаданные
        line_no=payload.line_no  # Указываем номер строки для суммы
    )

    # Помечаем строку счета как использованную если она существует
    if line:
        mark_invoice_line_used(db, payload.invoice_id, payload.line_no)

    db.commit()

    try:
        # Уведомляем об успешном применении счета
        websocket_manager.broadcast_to_batch(registry.imported_batch, {
            "type": "invoice_applied",
            "batch_id": registry.imported_batch,
            "registry_id": registry.id,
            "invoice_id": payload.invoice_id,
            "contractor": registry.contractor
        })
    except Exception as e:
        log.error(f"Failed to send WebSocket notification: {e}")

    log.info(
        "[APPLY_LINE] APPLIED OK registry_id=%s contractor=%s amount=%s position=%s",
        registry.id,
        registry.contractor,
        registry.amount,
        registry.position,
    )

    return {"status": "ok"}

# -------------------------------------------------------------------
# APPLY BATCH INVOICES
# -------------------------------------------------------------------

@app.post("/invoice/apply-batch/{batch_id}", summary="Применить OCR-счета ко всем строкам batch")
def apply_batch_invoices(batch_id: str, db: Session = Depends(get_db)):
    registries = (
        db.query(PaymentRegistry)
        .filter(PaymentRegistry.imported_batch == batch_id)
        .all()
    )

    if not registries:
        raise HTTPException(404, detail="Batch not found or empty")

    invoices = list_invoices(batch_id)

    for r in registries:
        if r.invoice_id:
            invoices.append({"id": r.invoice_id, "data": r.invoice_details})

    applied_count = 0

    for registry in registries:
        vehicle_text = (registry.vehicle or "").lower()
        license_plate_text = (registry.license_plate or "").replace(" ", "").lower()

        for invoice in invoices:
            data = invoice.get("data") or {}
            invoice_full_text = (data.get("invoice_full_text") or "").lower()

            if not invoice_full_text:
                continue

            if vehicle_text.split()[0] in invoice_full_text or license_plate_text in invoice_full_text:
                apply_invoice_ocr_to_registry(db, registry.id, invoice)
                applied_count += 1
                break

    db.commit()
    return {
        "status": "ok",
        "registries_processed": len(registries),
        "invoices_applied": applied_count,
    }

# -------------------------------------------------------------------
# UPLOAD EXCEL/PDF
# -------------------------------------------------------------------

@app.post("/upload", summary="Upload Excel or PDF")
async def upload_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    batch_id: str | None = Form(None),
    db: Session = Depends(get_db),
):
    ext = file.filename.split(".")[-1].lower()
    if ext not in ("xlsx", "xls", "pdf"):
        raise HTTPException(400, "Supported: xlsx, xls, pdf")

    if not batch_id:
        batch_id = str(uuid.uuid4())

    file_path = BASE_UPLOAD_DIR / f"{batch_id}_{file.filename}"

    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # ----------------------------------------------------------------
    # EXCEL
    # ----------------------------------------------------------------
    if ext in ("xlsx", "xls"):
        parser = ExcelParser()
        try:
            # Используем новый метод, который возвращает данные с позициями
            parsed_data = parser.parse_file_with_positions(file_path)
        except Exception as e:
            raise HTTPException(500, f"Excel parse error: {e}")

        grouped: dict[str, list[dict]] = defaultdict(list)

        # Группируем строки по номеру машины, сохраняя позиции
        for row_data, excel_position in parsed_data:
            plate = row_data.get("license_plate")
            if plate:
                grouped[plate].append({
                    "row_data": row_data,
                    "excel_position": excel_position
                })

        created_items = []
        current_position = 0
        
        for plate, rows in grouped.items():
            # Создаем ImportedRequest для каждой строки в группе
            for row_info in rows:
                create_imported_request(db, row_info["row_data"], batch_id, file.filename, ext)

            first = rows[0]
            comments = list(
                dict.fromkeys(
                    row_info["row_data"]["item_name"] for row_info in rows if row_info["row_data"].get("item_name")
                )
            )

            # Создаем строку реестра с сохранением позиции из Excel
            registry_item = create_payment_registry_item(
                db,
                {
                    "supplier": None,
                    "vehicle": first["row_data"].get("car_brand"),
                    "license_plate": plate,
                    "amount": 0,
                    "vat_amount": 0,
                    "comment": "; ".join(comments),
                    "matched_request_id": None,
                },
                batch_id,
                position=first["excel_position"]  # Передаем позицию из Excel 
            )
            created_items.append(registry_item)
            current_position += 1

        db.commit()

        registry_preview = build_registry_from_batch(db, batch_id)
    
        print(f"\n=== DEBUG /upload response ===")
        print(f"Excel parsed: {len(parsed_data)} rows")
        print(f"Registry preview: {len(registry_preview)} items")
        print(f"Positions: {[item.get('position') for item in registry_preview]}")
        
        # Возвращаем совместимый формат
        return {
            "message": f"Excel imported, {len(parsed_data)} rows processed",
            "batch_id": batch_id,
            "registry_preview": registry_preview,
            "data": registry_preview,  # Дублирование для совместимости
        }

    # ----------------------------------------------------------------
    # PDF
    # ----------------------------------------------------------------
    background_tasks.add_task(
        process_invoice_pdf_background,
        file_path,
        batch_id,
    )

    return {
        "message": "Invoice accepted for processing",
        "status": "processing",
        "batch_id": batch_id,
    }

# -------------------------------------------------------------------
# PREVIEW
# -------------------------------------------------------------------

@app.get("/invoice/{batch_id}/preview")
def get_invoice_preview(batch_id: str, db: Session = Depends(get_db)):
    print(f"\n=== DEBUG /invoice/{batch_id}/preview ===")
    
    # Получаем данные реестра из базы (сортировка по position)
    registry_preview = build_registry_from_batch(db, batch_id)
    
    print(f"Batch ID from URL: {batch_id}")
    print(f"Registry preview items: {len(registry_preview)}")
    print(f"Order by position: {[item.get('position') for item in registry_preview]}")
    
    if len(registry_preview) == 0:
        print(" WARNING: Registry preview is EMPTY!")
    
    # Получаем все счета для этого batch
    pending_invoices = list_invoices(batch_id)
    
    # Создаем словарь счетов по ID для быстрого доступа
    invoices_by_id = {}
    for inv in pending_invoices:
        inv_id = inv.get("id")
        if inv_id and inv.get("data"):
            data = inv["data"]
            
            # Создаем invoice_full_text если его нет
            invoice_full_text = data.get("invoice_full_text")
            if not invoice_full_text:
                invoice_number = data.get("invoice_number")
                invoice_date = data.get("invoice_date")
                if invoice_number and invoice_date:
                    invoice_full_text = f"Счет на оплату № {invoice_number} от {invoice_date}"
                elif invoice_number:
                    invoice_full_text = f"Счет на оплату № {invoice_number}"
                elif invoice_date:
                    invoice_full_text = f"Счет от {invoice_date}"
                else:
                    invoice_full_text = ""
            
            invoices_by_id[inv_id] = {
                "contractor": data.get("contractor"),
                "invoice_number": data.get("invoice_number"),
                "invoice_date": data.get("invoice_date"),
                "total": data.get("total"),
                "inn": data.get("inn"),
                "account": data.get("account"),
                "metadata_found": data.get("metadata_found", False),
                "invoice_full_text": invoice_full_text,
            }
    
    # Обновляем registry_preview данными из счетов
    updated_registry = []
    for item in registry_preview:
        invoice_id = item.get("invoice_id")
        
        # Если у строки есть invoice_id и у нас есть данные для этого счета
        if invoice_id and invoice_id in invoices_by_id:
            invoice_data = invoices_by_id[invoice_id]
            
            # Обновляем invoice_details в item
            current_details = item.get("invoice_details", {})
            if not isinstance(current_details, dict):
                current_details = {}
            
            # Объединяем существующие данные с новыми
            updated_details = {**current_details, **invoice_data}
            item["invoice_details"] = updated_details
            
            # Также обновляем contractor если его нет
            if not item.get("contractor") and invoice_data.get("contractor"):
                item["contractor"] = invoice_data.get("contractor")
        
        updated_registry.append(item)
    
    # Форматируем данные счетов для фронтенд
    invoices_data = []
    for inv in pending_invoices:
        invoice_data = {
            "id": inv.get("id"),
            "file": inv.get("file", ""),
            "confidence": inv.get("confidence", 0),
            "status": "pending",
        }
        
        if inv.get("data"):
            invoice_data["has_metadata"] = True
            invoice_data["ocr_data"] = inv["data"]
        else:
            invoice_data["has_metadata"] = False
        
        invoice_data["has_lines"] = bool(inv.get("lines") and len(inv["lines"]) > 0)
        invoices_data.append(invoice_data)
    
    print(f"Returning {len(updated_registry)} registry items and {len(pending_invoices)} invoices")
    
    response_data = {
        "registry_preview": updated_registry,
        "pending_invoices": len(pending_invoices),
        "invoices": invoices_data,
        "batch_id": batch_id,
    }
    
    return response_data

# -------------------------------------------------------------------
# UNMATCHED INVOICES
# -------------------------------------------------------------------

@app.get("/invoices/unmatched/{batch_id}")
def unmatched(batch_id: str):
    return list_invoices(batch_id)

# -------------------------------------------------------------------
# INVOICE LINES
# -------------------------------------------------------------------

@app.get("/invoice/{invoice_id}/lines")
def get_invoice_lines_endpoint(
    invoice_id: str,
    db: Session = Depends(get_db),
):
    print(f"\n=== DEBUG /invoice/{invoice_id}/lines ===")
    
    # 1. Проверяем что строка счета существует в буфере
    # Импорт внутри функции чтобы избежать циклического импорта
    from services.invoice_buffer import get_invoice
    
    invoice_in_buffer = get_invoice(invoice_id)
    print(f"Invoice in buffer: {invoice_in_buffer is not None}")
    if invoice_in_buffer:
        print(f"Buffer lines count: {len(invoice_in_buffer.get('lines', []))}")
    
    # 2. Проверяем строки в базе данных
    db_lines = (
        db.query(InvoiceLine)
        .filter(InvoiceLine.invoice_id == invoice_id)
        .order_by(InvoiceLine.line_no)
        .all()
    )
    
    print(f"Lines in database: {len(db_lines)}")
    for line in db_lines:
        print(f"  Line {line.line_no}: '{line.description[:30]}...' qty={line.quantity}, price={line.price}, total={line.total}")
    
    # 3. Если в базе нет строк, проверяем буфер
    result = []
    
    if db_lines:
        # Берем строки из базы
        for l in db_lines:
            try:
                line_data = {
                    "line_no": l.line_no,
                    "description": l.description,
                    "quantity": l.quantity,
                    "price": float(l.price) if l.price else None,
                    "total": float(l.total) if l.total else None,
                    "used": l.used,
                }
                result.append(line_data)
            except Exception as e:
                print(f"  Error formatting line {l.line_no}: {e}")
    elif invoice_in_buffer and 'lines' in invoice_in_buffer:
        # Берем строки из буфера
        buffer_lines = invoice_in_buffer.get('lines', [])
        print(f"Using {len(buffer_lines)} lines from buffer")
        
        for i, line in enumerate(buffer_lines):
            try:
                line_data = {
                    "line_no": line.get('line_no', i + 1),
                    "description": line.get('description', ''),
                    "quantity": int(line.get('qty', 1)),
                    "price": float(line.get('price', 0)) if line.get('price') else 0,
                    "total": float(line.get('total', 0)) if line.get('total') else 0,
                    "used": line.get('used', False),
                }
                result.append(line_data)
            except Exception as e:
                print(f"  Error formatting buffer line {i}: {e}")
    
    print(f"Returning {len(result)} lines to frontend")
    
    if not result:
        print("WARNING: No lines found for this invoice!")
        print(f"Invoice ID: {invoice_id}")
        print(f"Check if invoice was parsed correctly and lines were saved.")
    
    return result

# -------------------------------------------------------------------
# INVOICES FROM BUFFER
# -------------------------------------------------------------------

@app.get("/registry/{batch_id}/invoices-from-buffer")
def get_invoices_from_buffer(batch_id: str):
    """Получить все счета из буфера для данного batch (для фронтенда)"""
    try:
        # Получаем счета из буфера
        pending_invoices = list_invoices(batch_id)
        
        print(f"\n=== DEBUG /invoices-from-buffer for batch {batch_id} ===")
        print(f"Found {len(pending_invoices)} invoices in buffer")
        
        # Форматируем для фронтенда
        formatted_invoices = []
        for inv in pending_invoices:
            data = inv.get("data", {})
            
            # Создаем invoice_full_text если его нет
            invoice_full_text = data.get("invoice_full_text")
            if not invoice_full_text:
                invoice_number = data.get("invoice_number")
                invoice_date = data.get("invoice_date")
                if invoice_number and invoice_date:
                    invoice_full_text = f"Счет на оплату № {invoice_number} от {invoice_date}"
                elif invoice_number:
                    invoice_full_text = f"Счет на оплату № {invoice_number}"
                elif invoice_date:
                    invoice_full_text = f"Счет от {invoice_date}"
                else:
                    invoice_full_text = ""
            
            formatted_invoices.append({
                "id": inv.get("id"),
                "file": inv.get("file", ""),
                "contractor": data.get("contractor"),
                "invoice_number": data.get("invoice_number"),
                "invoice_date": data.get("invoice_date"),
                "total": data.get("total"),
                "invoice_full_text": invoice_full_text,
                "has_lines": bool(inv.get("lines") and len(inv["lines"]) > 0),
                "lines_count": len(inv.get("lines", [])),
                "status": "pending"
            })
        
        print(f"Returning {len(formatted_invoices)} formatted invoices")
        
        return {
            "invoices": formatted_invoices,
            "count": len(formatted_invoices),
            "batch_id": batch_id
        }
        
    except Exception as e:
        log.error(f"Error getting invoices from buffer: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# -------------------------------------------------------------------
# APPLY INVOICE METADATA
# -------------------------------------------------------------------

@app.post("/invoice/apply-metadata")
def apply_invoice_metadata(payload: ApplyInvoiceMetadataRequest, db: Session = Depends(get_db)):
    """Применить только метаданные счета без изменения суммы"""
    
    registry = db.query(PaymentRegistry).get(payload.registry_id)
    if not registry:
        raise HTTPException(404, "Registry item not found")

    invoice = get_invoice(payload.invoice_id)
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    invoice_data = {
        "id": payload.invoice_id,
        "data": dict(invoice["data"]),
        "confidence": invoice.get("confidence", 0)
    }

    # Создаем модифицированные данные с обнуленной суммой
    modified_data = invoice_data.copy()
    if "data" in modified_data:
        # Удаляем сумму из данных, если не хотим ее применять
        if "total" in modified_data["data"]:
            del modified_data["data"]["total"]
        if "amount" in modified_data["data"]:
            del modified_data["data"]["amount"]

    # Применяем с указанием не применять полные метаданные
    registry = apply_invoice_ocr_to_registry(
        db, 
        registry.id, 
        modified_data,
        apply_full_metadata=False  # Не применяем автоматически contractor и т.д.
    )
    
    # Вручную применяем только выбранные поля
    data = invoice["data"]
    
    if "contractor" in payload.apply_fields and data.get("contractor"):
        registry.contractor = data["contractor"]
    
    if "invoice_number" in payload.apply_fields and data.get("invoice_number"):
        registry.invoice_number = data["invoice_number"]
    
    if "invoice_date" in payload.apply_fields and data.get("invoice_date"):
        registry.invoice_date = data["invoice_date"]
    
    # Обновляем invoice_full_text если нужно
    if "invoice_full_text" in payload.apply_fields and data.get("invoice_full_text"):
        registry.invoice_full_text = data["invoice_full_text"]
    
    db.commit()
    
    return {"status": "ok", "applied_fields": payload.apply_fields}

# -------------------------------------------------------------------
# MANUAL INVOICE MATCH
# -------------------------------------------------------------------

@app.post("/invoice/manual-match")
def manual_invoice_match(payload: ManualInvoiceMatchRequest, db: Session = Depends(get_db)):
    """Ручное сопоставление счета со строкой реестра"""
    
    # 1. Проверяем существование записи реестра
    registry = db.query(PaymentRegistry).get(payload.registry_id)
    if not registry:
        raise HTTPException(404, "Registry item not found")
    
    # 2. Проверяем что запись принадлежит указанному batch
    if registry.imported_batch != payload.batch_id:
        raise HTTPException(400, "Registry item does not belong to this batch")
    
    # 3. Получаем счет из буфера
    invoice = get_invoice(payload.invoice_id)
    if not invoice:
        raise HTTPException(404, "Invoice not found")
    
    # 4. Подготавливаем данные счета
    invoice_data = {
        "id": payload.invoice_id,
        "data": dict(invoice["data"]),
        "lines": invoice.get("lines", []),
        "confidence": invoice.get("confidence", 0)
    }
    
    # 5. Применяем в зависимости от типа
    if payload.apply_type == "full":
        # Применяем все данные (метаданные + сумму из первой строки)
        apply_invoice_ocr_to_registry(
            db, registry.id, invoice_data,
            apply_full_metadata=True,
            line_no=0  # Первая строка счета
        )
        
    elif payload.apply_type == "metadata_only":
        # Применяем только метаданные
        apply_invoice_ocr_to_registry(
            db, registry.id, invoice_data,
            apply_full_metadata=True,
            line_no=None  # Не применяем сумму
        )
        
    elif payload.apply_type == "amount_only":
        # Применяем только сумму из первой строки
        apply_invoice_ocr_to_registry(
            db, registry.id, invoice_data,
            apply_full_metadata=False,  # Не применяем метаданные
            line_no=0  # Применяем сумму
        )
    
    db.commit()
    
    return {
        "status": "ok",
        "message": f"Invoice applied ({payload.apply_type})",
        "registry_id": registry.id,
        "invoice_id": payload.invoice_id,
        "contractor": registry.contractor,
        "amount": registry.amount,
        "position": registry.position
    }

# -------------------------------------------------------------------
# REORDER REGISTRY
# -------------------------------------------------------------------

@app.post("/registry/reorder")
def reorder_registry(request: ReorderRequest, db: Session = Depends(get_db)):
    """
    Изменение порядка строк в реестре.
    Принимает массив элементов с id и новой позицией.
    """
    print(f"\n=== DEBUG /registry/reorder for batch {request.batch_id} ===")
    
    updated_count = 0
    error_messages = []
    
    for item in request.items:
        try:
            registry_id = item.get("id")
            new_position = item.get("position")
            
            if not registry_id or new_position is None:
                error_messages.append(f"Invalid item: {item}")
                continue
            
            print(f"Updating registry_id {registry_id} to position {new_position}")
            
            registry = db.query(PaymentRegistry).filter(
                PaymentRegistry.id == registry_id,
                PaymentRegistry.imported_batch == request.batch_id
            ).first()
            
            if not registry:
                error_messages.append(f"Registry item {registry_id} not found in batch {request.batch_id}")
                continue
            
            # Обновляем позицию
            old_position = registry.position
            registry.position = new_position
            updated_count += 1
            
            print(f"Updated: id={registry_id}, old={old_position}, new={new_position}")
            
        except Exception as e:
            error_messages.append(f"Error updating registry_id {registry_id}: {str(e)}")
    
    db.commit()
    
    # Получаем обновленный порядок
    registry_preview = build_registry_from_batch(db, request.batch_id)
    
    response = {
        "status": "ok",
        "message": f"Updated {updated_count} items",
        "updated_count": updated_count,
        "registry_preview": registry_preview
    }
    
    if error_messages:
        response["errors"] = error_messages
    
    return response

# -------------------------------------------------------------------
# GET REGISTRY ORDER
# -------------------------------------------------------------------

@app.get("/registry/{batch_id}/order")
def get_registry_order(batch_id: str, db: Session = Depends(get_db)):
    """
    Получить текущий порядок строк реестра.
    Возвращает mapping id -> position.
    """
    registries = (
        db.query(PaymentRegistry.id, PaymentRegistry.position)
        .filter(PaymentRegistry.imported_batch == batch_id)
        .order_by(PaymentRegistry.position)
        .all()
    )
    
    order_map = {registry.id: registry.position for registry in registries}
    
    return {
        "batch_id": batch_id,
        "order": order_map,
        "total_items": len(registries)
    }

# -------------------------------------------------------------------
# ROOT
# -------------------------------------------------------------------

@app.get("/")
def root():
    return {"message": "Registry Control API"}

# -------------------------------------------------------------------
# DEBUG ENDPOINTS
# -------------------------------------------------------------------

@app.get("/debug/invoice/{invoice_id}")
def debug_invoice_info(invoice_id: str, db: Session = Depends(get_db)):
    """Отладочная информация о счете"""
    
    # Проверяем буфер
    from services.invoice_buffer import get_invoice
    
    invoice_in_buffer = get_invoice(invoice_id)
    
    # Проверяем базу
    db_lines = db.query(InvoiceLine).filter(InvoiceLine.invoice_id == invoice_id).all()
    
    return {
        "invoice_id": invoice_id,
        "in_buffer": invoice_in_buffer is not None,
        "buffer_data": {
            "has_data": invoice_in_buffer is not None,
            "has_lines": invoice_in_buffer and 'lines' in invoice_in_buffer,
            "lines_count": len(invoice_in_buffer.get('lines', [])) if invoice_in_buffer else 0,
            "contractor": invoice_in_buffer.get('data', {}).get('contractor') if invoice_in_buffer else None,
            "metadata_keys": list(invoice_in_buffer.get('data', {}).keys()) if invoice_in_buffer else []
        },
        "in_database": {
            "lines_count": len(db_lines),
            "lines": [
                {
                    "line_no": l.line_no,
                    "description": l.description[:50],
                    "quantity": l.quantity,
                    "price": l.price,
                    "total": l.total
                }
                for l in db_lines[:5]  # Первые 5 строк
            ]
        }
    }


@app.get("/debug/check-invoice/{invoice_id}")
def debug_check_invoice(invoice_id: str, db: Session = Depends(get_db)):
    """Отладочная информация о счете"""
    from services.invoice_buffer import get_invoice
    
    # 1. Проверяем буфер
    invoice_in_buffer = get_invoice(invoice_id)
    
    # 2. Проверяем базу
    db_lines = db.query(InvoiceLine).filter(InvoiceLine.invoice_id == invoice_id).all()
    
    # 3. Форматируем информацию
    buffer_info = {}
    if invoice_in_buffer:
        buffer_info = {
            "id": invoice_in_buffer.get("id"),
            "batch_id": invoice_in_buffer.get("batch_id"),
            "has_lines": "lines" in invoice_in_buffer,
            "lines_count": len(invoice_in_buffer.get("lines", [])),
            "lines_sample": invoice_in_buffer.get("lines", [])[:2] if invoice_in_buffer.get("lines") else [],
            "metadata": invoice_in_buffer.get("data", {})
        }
    
    return {
        "invoice_id": invoice_id,
        "in_buffer": buffer_info,
        "in_database": {
            "lines_count": len(db_lines),
            "lines": [
                {
                    "line_no": l.line_no,
                    "description": l.description[:50] if l.description else "",
                    "quantity": l.quantity,
                    "price": l.price,
                    "total": l.total
                }
                for l in db_lines[:5]
            ]
        },
        "api_endpoints": {
            "lines": f"/invoice/{invoice_id}/lines",
            "buffer": f"/debug/invoice/{invoice_id}"
        }
    }

# -------------------------------------------------------------------
# MULTIPLE LINES APPLICATION
# -------------------------------------------------------------------

@app.post("/invoice/apply-multiple-lines")
def apply_multiple_invoice_lines(payload: ApplyMultipleLinesRequest, db: Session = Depends(get_db)):
    """Применить несколько выбранных строк счета к строке реестра"""
    log.info(
        "[APPLY_MULTIPLE] invoice_id=%s lines=%s registry_id=%s",
        payload.invoice_id,
        payload.line_nos,
        payload.registry_id,
    )

    registry = db.query(PaymentRegistry).get(payload.registry_id)
    if not registry:
        log.warning("[APPLY_MULTIPLE] registry not found")
        raise HTTPException(404, "Registry item not found")

    # Проверяем что запись принадлежит указанному batch
    if registry.imported_batch != payload.batch_id:
        raise HTTPException(400, "Registry item does not belong to this batch")

    # Получаем данные счета из буфера
    invoice = get_invoice(payload.invoice_id)
    if not invoice:
        log.error("[APPLY_MULTIPLE] invoice not found in buffer")
        raise HTTPException(404, "Invoice not found")

    # Суммируем суммы выбранных строк
    total_amount = 0
    applied_lines = []
    
    for line_no in payload.line_nos:
        # Ищем строку в базе данных
        line = (
            db.query(InvoiceLine)
            .filter(
                InvoiceLine.invoice_id == payload.invoice_id,
                InvoiceLine.line_no == line_no,
                InvoiceLine.used.is_(False),
            )
            .first()
        )
        
        if line:
            try:
                line_total = float(line.total) if line.total else 0
                total_amount += line_total
                applied_lines.append(line_no)
                
                # Помечаем строку как использованную
                mark_invoice_line_used(db, payload.invoice_id, line_no)
            except (ValueError, TypeError):
                log.warning(f"Invalid total for line {line_no}: {line.total}")
        else:
            log.warning(f"Line {line_no} not found or already used")

    log.info(f"[APPLY_MULTIPLE] Total amount: {total_amount} from {len(applied_lines)} lines")

    invoice_data = {
        "id": payload.invoice_id,
        "data": dict(invoice["data"]),
        "lines": invoice.get("lines", []),
        "confidence": invoice.get("confidence", 0)
    }

    # Создаем модифицированные данные с суммой выбранных строк
    modified_data = invoice_data.copy()
    if "data" in modified_data:
        # Добавляем общую сумму выбранных строк
        modified_data["data"]["total"] = str(total_amount)
    
    # Применяем счет к строке реестра
    apply_invoice_ocr_to_registry(
        db, 
        registry.id, 
        modified_data,
        apply_full_metadata=True,
        line_no=None  # Не используем конкретную строку, так как сумма уже вычислена
    )

    # Обновляем сумму в реестре суммой выбранных строк
    registry.amount = total_amount

    db.commit()

    log.info(
        "[APPLY_MULTIPLE] APPLIED OK registry_id=%s total_amount=%s lines_applied=%s",
        registry.id,
        total_amount,
        len(applied_lines),
    )

    return {
        "status": "ok",
        "message": f"Applied {len(applied_lines)} lines",
        "total_amount": total_amount,
        "lines_applied": applied_lines,
        "registry_id": registry.id
    }

@app.post("/invoice/apply-all-lines")
def apply_all_invoice_lines(payload: ApplyAllLinesRequest, db: Session = Depends(get_db)):
    """Применить все строки счета к строке реестра"""
    log.info(
        "[APPLY_ALL] invoice_id=%s registry_id=%s",
        payload.invoice_id,
        payload.registry_id,
    )

    registry = db.query(PaymentRegistry).get(payload.registry_id)
    if not registry:
        log.warning("[APPLY_ALL] registry not found")
        raise HTTPException(404, "Registry item not found")

    # Проверяем что запись принадлежит указанному batch
    if registry.imported_batch != payload.batch_id:
        raise HTTPException(400, "Registry item does not belong to this batch")

    # Получаем данные счета из буфера
    invoice = get_invoice(payload.invoice_id)
    if not invoice:
        log.error("[APPLY_ALL] invoice not found in buffer")
        raise HTTPException(404, "Invoice not found")

    # Получаем все строки счета
    lines = (
        db.query(InvoiceLine)
        .filter(
            InvoiceLine.invoice_id == payload.invoice_id,
            InvoiceLine.used.is_(False),
        )
        .all()
    )

    # Суммируем все суммы строк
    total_amount = 0
    applied_lines = []
    
    for line in lines:
        try:
            line_total = float(line.total) if line.total else 0
            total_amount += line_total
            applied_lines.append(line.line_no)
            
            # Помечаем строку как использованную
            mark_invoice_line_used(db, payload.invoice_id, line.line_no)
        except (ValueError, TypeError):
            log.warning(f"Invalid total for line {line.line_no}: {line.total}")

    log.info(f"[APPLY_ALL] Total amount: {total_amount} from {len(applied_lines)} lines")

    invoice_data = {
        "id": payload.invoice_id,
        "data": dict(invoice["data"]),
        "lines": invoice.get("lines", []),
        "confidence": invoice.get("confidence", 0)
    }

    # Создаем модифицированные данные с общей суммой
    modified_data = invoice_data.copy()
    if "data" in modified_data:
        # Добавляем общую сумму всех строк
        modified_data["data"]["total"] = str(total_amount)
    
    # Применяем счет к строке реестра
    apply_invoice_ocr_to_registry(
        db, 
        registry.id, 
        modified_data,
        apply_full_metadata=True,
        line_no=None  # Используем общую сумму
    )

    # Обновляем сумму в реестре общей суммой всех строк
    registry.amount = total_amount

    db.commit()

    log.info(
        "[APPLY_ALL] APPLIED OK registry_id=%s total_amount=%s lines_applied=%s",
        registry.id,
        total_amount,
        len(applied_lines),
    )

    return {
        "status": "ok",
        "message": f"Applied all {len(applied_lines)} lines",
        "total_amount": total_amount,
        "lines_applied": applied_lines,
        "registry_id": registry.id
    }


# -------------------------------------------------------------------
# DEFECT SHEET MODULE (Дефектная ведомость)
# -------------------------------------------------------------------

UPLOAD_DIR = Path("uploads/defect_sheets")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


@app.post("/api/defect/upload", summary="Загрузить дефектную ведомость (Excel)")
async def upload_defect_sheet(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    batch_id: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    Загружает Excel файл дефектной ведомости.
    Возвращает batch_id для дальнейшей работы.
    """
    # Проверка расширения
    ext = file.filename.split(".")[-1].lower()
    if ext not in ("xlsx", "xls"):
        raise HTTPException(400, "Поддерживаются только Excel файлы (.xlsx, .xls)")
    
    # Генерируем batch_id если не передан
    if not batch_id:
        batch_id = str(uuid.uuid4())
    
    # Сохраняем файл
    file_path = UPLOAD_DIR / f"{batch_id}_{file.filename}"
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    # Запускаем фоновую задачу парсинга
    background_tasks.add_task(
        process_defect_sheet_background,
        file_path,
        batch_id,
        file.filename
    )
    
    # Сразу создаем запись в БД со статусом processing
    sheet = create_defect_sheet(db, file.filename, batch_id)
    db.commit()
    
    # Уведомляем через WebSocket о начале обработки
    await websocket_manager.broadcast_to_batch(batch_id, {
        "type": "defect_sheet_processing",
        "batch_id": batch_id,
        "sheet_id": sheet.id,
        "status": "processing",
        "message": "Файл принят, начинаем парсинг"
    })
    
    return {
        "message": "Файл принят в обработку",
        "batch_id": batch_id,
        "sheet_id": sheet.id,
        "status": "processing"
    }


def process_defect_sheet_background(file_path: Path, batch_id: str, original_filename: str):
    """Фоновая задача для парсинга дефектной ведомости"""
    from crud import create_defect_sheet, create_defect_sheet_items, get_defect_sheet_by_batch
    from datetime import datetime
    import asyncio
    
    db = SessionLocal()
    sheet = None
    items = []
    metadata = {}
    
    try:
        log.info(f"[DEFECT] Starting parsing of {file_path.name}")
        
        # Обновляем статус через WebSocket
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(
                websocket_manager.broadcast_to_batch(batch_id, {
                    "type": "defect_sheet_status",
                    "batch_id": batch_id,
                    "status": "parsing",
                    "progress": 30
                })
            )
            loop.close()
        except Exception as ws_error:
            log.warning(f"WebSocket status update failed: {ws_error}")
        
        # Парсим файл
        try:
            items, metadata = parse_defect_sheet(file_path)
            log.info(f"[DEFECT] Parsed {len(items)} items")
        except Exception as parse_error:
            log.exception(f"[DEFECT] Parse error: {parse_error}")
            items = []
            metadata = {"error": str(parse_error)}
        
        # Находим запись в БД или создаем новую
        sheet = get_defect_sheet_by_batch(db, batch_id)
        if not sheet:
            log.info(f"[DEFECT] Sheet not found for batch {batch_id}, creating new one")
            sheet = create_defect_sheet(db, batch_id, original_filename)
            log.info(f"[DEFECT] Created new sheet with id {sheet.id}")
        
        # Обновляем метаданные
        if metadata.get("period_start"):
            sheet.period_start = metadata["period_start"]
        if metadata.get("period_end"):
            sheet.period_end = metadata["period_end"]
        
        # Сохраняем строки (если есть)
        if items:
            try:
                create_defect_sheet_items(db, sheet.id, items)
                log.info(f"[DEFECT] Saved {len(items)} items to DB")
            except Exception as db_error:
                log.exception(f"[DEFECT] DB save error: {db_error}")
        
        # Обновляем статус
        sheet.status = "processed" if items else "no_data"
        sheet.total_items = len(items)
        db.commit()
        
        # НЕБОЛЬШАЯ ЗАДЕРЖКА перед отправкой WebSocket
        import time
        time.sleep(1)  # Даем время клиенту подключиться
        
        # Отправляем WebSocket сообщение
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            message = {
                "type": "defect_sheet_processed",
                "batch_id": batch_id,
                "sheet_id": sheet.id,
                "status": "processed",
                "total_items": len(items),
                "metadata": metadata
            }
            
            log.info(f"📤 Sending WebSocket message: {message}")
            loop.run_until_complete(
                websocket_manager.broadcast_to_batch(batch_id, message)
            )
            loop.close()
            
        except Exception as ws_error:
            log.error(f"Failed to send WebSocket notification: {ws_error}")
        
        log.info(f"[DEFECT] Successfully processed {file_path.name}")
        
    except Exception as e:
        log.exception(f"[DEFECT] Error processing {file_path.name}: {e}")
        
        # Обновляем статус на ошибку
        if sheet:
            sheet.status = "error"
            db.commit()
        
        # Отправляем сообщение об ошибке
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(
                websocket_manager.broadcast_to_batch(batch_id, {
                    "type": "defect_sheet_error",
                    "batch_id": batch_id,
                    "error": str(e)
                })
            )
            loop.close()
        except:
            pass
    finally:
        db.close()


@app.get("/api/defect/{batch_id}/preview", response_model=DefectSheetPreviewResponse)
def preview_defect_sheet(batch_id: str, db: Session = Depends(get_db)):
    """
    Получить предпросмотр дефектной ведомости.
    Возвращает все распарсенные строки.
    """
    sheet = get_defect_sheet_by_batch(db, batch_id)
    if not sheet:
        raise HTTPException(404, f"Дефектная ведомость с batch_id {batch_id} не найдена")
    
    items = get_defect_sheet_items(db, sheet.id)
    
    return {
        "sheet_id": sheet.id,
        "batch_id": sheet.batch_id,
        "file_name": sheet.file_name,
        "upload_date": sheet.upload_date,
        "status": sheet.status,
        "period_start": sheet.period_start,
        "period_end": sheet.period_end,
        "total_items": len(items),
        "items": items
    }


@app.get("/api/defect/{sheet_id}/items")
def get_defect_items(sheet_id: int, db: Session = Depends(get_db)):
    """Получить все строки дефектной ведомости по ID"""
    items = get_defect_sheet_items(db, sheet_id)
    
    print(f"📊 Found {len(items)} items in DB for sheet {sheet_id}")  # Отладка

    # Преобразуем SQLAlchemy объекты в словари
    result = []
    for item in items:
        result.append({
            "id": item.id,
            "position": item.position,
            "excel_position": item.excel_position,
            "subposition": item.subposition,
            "address": item.address,
            "material_name": item.material_name,
            "requested_quantity": float(item.requested_quantity) if item.requested_quantity else None,
            "weight_tons": float(item.weight_tons) if item.weight_tons else None,
            "calculated_meters": float(item.calculated_meters) if item.calculated_meters else None,
            "profile_type": item.profile_type,
            "profile_params": item.profile_params,
            "formula_used": item.formula_used,
            "is_calculated": item.is_calculated,
            "selected_for_calculation": item.selected_for_calculation,
            "calculated_at": item.calculated_at.isoformat() if item.calculated_at else None,
            "requirement_number": item.requirement_number,
            "requirement_date": item.requirement_date.isoformat() if item.requirement_date else None,
            "car_brand": item.car_brand,
            "license_plate": item.license_plate,
            "recipient": item.recipient,
            "article": item.article
        })
    
    # Возвращаем объект с полем items (как ожидает фронтенд)
    return {
        "items": result,
        "total": len(result)
    }


@app.post("/api/defect/calculate")
def calculate_defect_items(
    request: CalculationRequest,
    db: Session = Depends(get_db)
):
    """
    Пересчет выбранных строк из тонн в метры.
    Если item_ids не указан или пуст - пересчитываются все непосчитанные.
    """
    sheet = get_defect_sheet(db, request.sheet_id)
    if not sheet:
        raise HTTPException(404, "Дефектная ведомость не найдена")
    
    # Получаем все строки
    all_items = get_defect_sheet_items(db, request.sheet_id)
    
    # Фильтруем строки для пересчета
    items_to_calc = []
    if request.item_ids:
        items_to_calc = [item for item in all_items if item.id in request.item_ids]
    else:
        items_to_calc = [item for item in all_items if not item.is_calculated]
    
    if not items_to_calc:
        return {
            "sheet_id": request.sheet_id,
            "total_items": len(all_items),
            "calculated_items": 0,
            "results": []
        }
    
    # Подготавливаем данные для калькулятора
    calc_items = []
    for item in items_to_calc:
        # Используем вес из строки или запрошенное количество
        weight = item.weight_tons or item.requested_quantity
        if not weight:
            continue
        
        # Определяем параметры профиля
        profile_params = request.profile_params.copy()
        if item.profile_params:
            profile_params.update(item.profile_params)
        
        calc_items.append({
            "id": item.id,
            "weight_tons": float(weight),
            "profile_type": request.profile_type,
            "profile_params": profile_params
        })
    
    # Пакетный пересчет
    results = metal_calculator.calculate_batch(
        items=calc_items,
        default_profile_type=request.profile_type,
        default_params=request.profile_params
    )
    
    # Сохраняем результаты в БД
    calculated_count = 0
    for result in results:
        if result.get("calculated_meters") and not result.get("error"):
            update_defect_sheet_item_calculation(
                db,
                result["id"],
                result["calculated_meters"],
                result["formula_used"]
            )
            calculated_count += 1
    
    # Обновляем статус ведомости
    if calculated_count > 0:
        all_calculated = all(item.is_calculated for item in all_items)
        sheet.status = "calculated" if all_calculated else "partially_calculated"
    
    db.commit()
    
    # Уведомляем через WebSocket
    asyncio.run(websocket_manager.broadcast_to_batch(sheet.batch_id, {
        "type": "defect_calculation_complete",
        "sheet_id": sheet.id,
        "batch_id": sheet.batch_id,
        "calculated_items": calculated_count,
        "total_items": len(all_items)
    }))
    
    return {
        "sheet_id": request.sheet_id,
        "total_items": len(all_items),
        "calculated_items": calculated_count,
        "results": results
    }


@app.get("/api/defect/formulas")
def get_calculation_formulas():
    """Получить список доступных формул для пересчета"""
    formulas = []
    for profile_type, info in metal_calculator.formulas.items():
        formulas.append({
            "type": profile_type,
            "name": info["name"],
            "formula": info["formula"],
            "description": info["description"],
            "params": info["params"]
        })
    return {"formulas": formulas}


@app.post("/api/defect/save")
def save_defect_sheet(
    sheet_id: int,
    db: Session = Depends(get_db)
):
    """
    Сохранить результаты пересчета (финализировать ведомость)
    """
    sheet = get_defect_sheet(db, sheet_id)
    if not sheet:
        raise HTTPException(404, "Дефектная ведомость не найдена")
    
    # Проверяем, что все строки обработаны
    items = get_defect_sheet_items(db, sheet_id)
    uncounted = [item for item in items if not item.is_calculated and item.weight_tons]
    
    if uncounted and sheet.status != "partially_calculated":
        return {
            "warning": f"Есть непересчитанные строки ({len(uncounted)}). Сохранить как есть?",
            "can_save": True
        }
    
    sheet.status = "exported"
    db.commit()
    
    # Уведомление
    asyncio.run(websocket_manager.broadcast_to_batch(sheet.batch_id, {
        "type": "defect_sheet_saved",
        "sheet_id": sheet.id,
        "batch_id": sheet.batch_id
    }))
    
    return {
        "status": "saved",
        "sheet_id": sheet_id,
        "message": "Дефектная ведомость сохранена"
    }

@app.post("/api/defect/export")
async def export_defect_sheet(
    request: ExportRequest,
    db: Session = Depends(get_db)
):
    """
    Экспорт дефектной ведомости в Excel (старый формат .xls)
    """
    sheet = get_defect_sheet(db, request.sheet_id)
    if not sheet:
        raise HTTPException(404, "Дефектная ведомость не найдена")
    
    items = get_defect_sheet_items(db, request.sheet_id)
    
    print(f"📊 Exporting {len(items)} items to .xls format")
    
    import xlwt
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    
    # Создаем workbook
    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Дефектная ведомость')
    
    # Стили
    header_style = xlwt.XFStyle()
    header_font = xlwt.Font()
    header_font.bold = True
    header_font.colour_index = 1  # белый
    header_font.height = 200  # 10pt * 20
    
    header_pattern = xlwt.Pattern()
    header_pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    header_pattern.pattern_fore_colour = 5  # синий
    
    header_style.font = header_font
    header_style.pattern = header_pattern
    header_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
    
    # Заголовки
    headers = [
        "№ п/п", "Марка (Адрес)", "Наименование материала",
        "Затреб (тонн)", "Вес (тонн)", "Тип профиля",
        "Параметры", "Пересчитано (метров)", "Формула", "Статус"
    ]
    
    # Записываем заголовки
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
        # Устанавливаем ширину колонки (1 единица = 1/256 символа)
        worksheet.col(col).width = 256 * 20  # 20 символов
    
    # Увеличиваем ширину для длинных колонок
    worksheet.col(2).width = 256 * 50  # Наименование материала
    worksheet.col(1).width = 256 * 25  # Марка
    worksheet.col(8).width = 256 * 30  # Формула
    
    # Стиль для чисел
    number_style = xlwt.XFStyle()
    number_style.num_format_str = '#,##0.000'
    number_style.alignment.horz = xlwt.Alignment.HORZ_RIGHT
    
    number_style_2dec = xlwt.XFStyle()
    number_style_2dec.num_format_str = '#,##0.00'
    number_style_2dec.alignment.horz = xlwt.Alignment.HORZ_RIGHT
    
    # Стиль для текста
    text_style = xlwt.XFStyle()
    text_style.alignment.horz = xlwt.Alignment.HORZ_LEFT
    
    # Стиль для статуса
    status_style = xlwt.XFStyle()
    status_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
    
    # Записываем данные
    row_count = 0
    for row, item in enumerate(items, start=1):
        try:
            # Основные поля
            worksheet.write(row, 0, item.position or "", text_style)
            worksheet.write(row, 1, item.address or "", text_style)
            worksheet.write(row, 2, item.material_name or "", text_style)
            
            # Числовые значения
            if item.requested_quantity:
                worksheet.write(row, 3, float(item.requested_quantity), number_style)
            else:
                worksheet.write(row, 3, "", text_style)
                
            if item.weight_tons:
                worksheet.write(row, 4, float(item.weight_tons), number_style)
            else:
                worksheet.write(row, 4, "", text_style)
            
            worksheet.write(row, 5, item.profile_type or "", text_style)
            worksheet.write(row, 6, str(item.profile_params) if item.profile_params else "", text_style)
            
            if item.calculated_meters:
                worksheet.write(row, 7, float(item.calculated_meters), number_style_2dec)
            else:
                worksheet.write(row, 7, "", text_style)
            
            worksheet.write(row, 8, item.formula_used or "", text_style)
            
            status = "✓ Пересчитано" if item.is_calculated else "Ожидает"
            worksheet.write(row, 9, status, status_style)
            
            row_count += 1
            
        except Exception as e:
            print(f"❌ Error writing row {row}: {e}")
            continue
    
    print(f"✅ Rows written: {row_count}")
    
    # Сохраняем в буфер
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    file_size = len(output.getvalue())
    print(f"✅ Excel .xls file created: {file_size} bytes")
    
    filename = f"defect_sheet_{sheet.batch_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.ms-excel",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@app.post("/api/defect/export-excel")
async def export_defect_sheet_excel(
    request: ExportRequest,
    db: Session = Depends(get_db)
):
    """
    Экспорт дефектной ведомости в форматированный Excel
    """
    sheet = get_defect_sheet(db, request.sheet_id)
    if not sheet:
        raise HTTPException(404, "Дефектная ведомость не найдена")
    
    items = get_defect_sheet_items(db, request.sheet_id)
    
    print(f"📊 Exporting {len(items)} items to formatted Excel")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    
    # Создаем workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Дефектная ведомость"
    
    # Стили
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    
    # Заливка для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Выравнивание
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовок документа
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "Дефектная ведомость"
    cell.font = title_font
    cell.alignment = center_alignment
    
    # Период
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = f"За период: {datetime.now().strftime('%d.%m.%Y')}"
    cell.alignment = center_alignment
    
    # Пустая строка
    ws.row_dimensions[3].height = 15
    
    # Заголовки таблицы
    headers = ["№ п/п", "Дата требования", "Марка/Адрес", "Наименование работ", "Наименование материалов"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Сортируем все items по позиции для сохранения порядка
    sorted_items = sorted(items, key=lambda x: x.position if x.position else 0)
    
    current_row = 5
    position_counter = 1
    
    # Записываем все строки подряд, без группировки по требованиям
    for idx, item in enumerate(sorted_items):
        # № п/п - сквозная нумерация
        cell = ws.cell(row=current_row, column=1, value=position_counter)
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Дата требования
        date_value = ""
        if item.requirement_date:
            if isinstance(item.requirement_date, datetime):
                date_value = item.requirement_date.strftime('%d.%m.%Y')
            else:
                date_value = str(item.requirement_date)
        
        cell = ws.cell(row=current_row, column=2, value=date_value)
        cell.border = thin_border
        cell.alignment = center_alignment
        
        # Марка/Адрес
        cell = ws.cell(row=current_row, column=3, value=item.address or "")
        cell.border = thin_border
        cell.alignment = left_alignment
        
        # Наименование работ (копируем адрес как в примере)
        cell = ws.cell(row=current_row, column=4, value=item.address or "")
        cell.border = thin_border
        cell.alignment = left_alignment
        
        #  ОБЪЕДИНЯЕМ ДАННЫЕ В КОЛОНКЕ НАИМЕНОВАНИЕ МАТЕРИАЛОВ 
        material_parts = []
        
        # Добавляем наименование материала
        if item.material_name:
            material_parts.append(item.material_name)
        
        # Добавляем пересчитанные метры, если есть
        if item.calculated_meters:
            material_parts.append(f"пересчитано: {float(item.calculated_meters):.2f} м")
        
        # Добавляем вес в тоннах, если есть
        if item.weight_tons:
            material_parts.append(f"вес: {float(item.weight_tons):.3f} т")
        
        # Объединяем все части через перенос строки
        combined_material = "\n".join(material_parts)
        
        cell = ws.cell(row=current_row, column=5, value=combined_material)
        cell.border = thin_border
        cell.alignment = left_alignment
        
        current_row += 1
        position_counter += 1
    
    # Устанавливаем ширину колонок
    ws.column_dimensions['A'].width = 10   # № п/п
    ws.column_dimensions['B'].width = 15   # Дата требования
    ws.column_dimensions['C'].width = 25   # Марка/Адрес
    ws.column_dimensions['D'].width = 30   # Наименование работ
    ws.column_dimensions['E'].width = 60   # Для объединенных данных
    
    # Добавляем границы для всей таблицы
    for row in ws.iter_rows(min_row=4, max_row=current_row-1, min_col=1, max_col=5):
        for cell in row:
            if not cell.border:
                cell.border = thin_border
    
    # Сохраняем в буфер
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"defect_sheet_{sheet.batch_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )



# добавление созданных строк в базу
class CreateDefectItemRequest(BaseModel):
    sheet_id: int
    position: Optional[int] = None
    address: Optional[str] = None
    material_name: Optional[str] = None
    requested_quantity: Optional[float] = None
    weight_tons: Optional[float] = None
    profile_type: Optional[str] = None
    profile_params: Optional[Dict[str, Any]] = None
    calculated_meters: Optional[float] = None
    formula_used: Optional[str] = None
    is_calculated: bool = False

@app.post("/api/defect/items", response_model=DefectSheetItemResponse)
async def create_defect_item(
    request: CreateDefectItemRequest,
    db: Session = Depends(get_db)
):
    """
    Создать новую строку в дефектной ведомости
    """
    # Проверяем существование ведомости
    sheet = db.query(DefectSheet).filter(DefectSheet.id == request.sheet_id).first()
    if not sheet:
        raise HTTPException(404, "Дефектная ведомость не найдена")
    
    # Определяем следующую позицию, если не указана
    position = request.position
    if position is None:
        max_pos = db.query(func.max(DefectSheetItem.position)).filter(
            DefectSheetItem.sheet_id == request.sheet_id
        ).scalar() or 0
        position = max_pos + 1
    
    # Создаем новую строку
    new_item = DefectSheetItem(
        sheet_id=request.sheet_id,
        position=position,
        address=request.address,
        material_name=request.material_name,
        requested_quantity=request.requested_quantity,
        weight_tons=request.weight_tons,
        profile_type=request.profile_type,
        profile_params=request.profile_params,
        calculated_meters=request.calculated_meters,
        formula_used=request.formula_used,
        is_calculated=request.is_calculated,
        selected_for_calculation=False
    )
    
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    
    # Обновляем общее количество в ведомости
    total_items = db.query(func.count(DefectSheetItem.id)).filter(
        DefectSheetItem.sheet_id == request.sheet_id
    ).scalar()
    sheet.total_items = total_items
    db.commit()
    
    # Отправляем WebSocket уведомление
    asyncio.create_task(websocket_manager.broadcast_to_batch(sheet.batch_id, {
        "type": "defect_item_created",
        "sheet_id": sheet.id,
        "batch_id": sheet.batch_id,
        "item_id": new_item.id,
        "position": new_item.position
    }))
    
    return new_item

# эндпоинт для удаления строки
@app.delete("/api/defect/items/{item_id}")
async def delete_defect_item(
    item_id: int,
    db: Session = Depends(get_db)
):
    """
    Удалить строку из дефектной ведомости
    """
    item = db.query(DefectSheetItem).filter(
        DefectSheetItem.id == item_id
    ).first()
    
    if not item:
        raise HTTPException(404, "Строка не найдена")
    
    sheet_id = item.sheet_id
    batch_id = item.sheet.batch_id
    
    db.delete(item)
    db.commit()
    
    # Обновляем общее количество
    total_items = db.query(func.count(DefectSheetItem.id)).filter(
        DefectSheetItem.sheet_id == sheet_id
    ).scalar()
    sheet = db.query(DefectSheet).filter(DefectSheet.id == sheet_id).first()
    if sheet:
        sheet.total_items = total_items
        db.commit()
    
    # Отправляем WebSocket уведомление
    asyncio.create_task(websocket_manager.broadcast_to_batch(batch_id, {
        "type": "defect_item_deleted",
        "sheet_id": sheet_id,
        "batch_id": batch_id,
        "item_id": item_id
    }))
    
    return {"status": "deleted", "item_id": item_id}


@app.delete("/api/defect/{sheet_id}")
def delete_defect_sheet_endpoint(
    sheet_id: int,
    db: Session = Depends(get_db)
):
    """Удалить дефектную ведомость"""
    sheet = get_defect_sheet(db, sheet_id)
    if not sheet:
        raise HTTPException(404, "Дефектная ведомость не найдена")
    
    batch_id = sheet.batch_id
    delete_defect_sheet(db, sheet_id)
    db.commit()
    
    return {"status": "deleted", "batch_id": batch_id}


# для более эффективного удаления строк batch delete

class BatchDeleteRequest(BaseModel):
    item_ids: List[int]

@app.post("/api/defect/items/batch-delete")
async def batch_delete_defect_items(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db)
):
    """
    Удалить несколько строк из дефектной ведомости
    """
    if not request.item_ids:
        raise HTTPException(400, "Список ID пуст")
    
    # Получаем первую строку для информации о sheet и batch
    first_item = db.query(DefectSheetItem).filter(
        DefectSheetItem.id == request.item_ids[0]
    ).first()
    
    if not first_item:
        raise HTTPException(404, "Строки не найдены")
    
    sheet_id = first_item.sheet_id
    batch_id = first_item.sheet.batch_id
    
    # Удаляем все указанные строки
    deleted = db.query(DefectSheetItem).filter(
        DefectSheetItem.id.in_(request.item_ids)
    ).delete(synchronize_session=False)
    
    db.commit()
    
    # Обновляем общее количество
    total_items = db.query(func.count(DefectSheetItem.id)).filter(
        DefectSheetItem.sheet_id == sheet_id
    ).scalar()
    
    sheet = db.query(DefectSheet).filter(DefectSheet.id == sheet_id).first()
    if sheet:
        sheet.total_items = total_items
        db.commit()
    
    # Отправляем WebSocket уведомление
    asyncio.create_task(websocket_manager.broadcast_to_batch(batch_id, {
        "type": "defect_items_deleted",
        "sheet_id": sheet_id,
        "batch_id": batch_id,
        "deleted_count": deleted
    }))
    
    return {
        "status": "deleted",
        "deleted_count": deleted,
        "item_ids": request.item_ids
    }

# -------------------------------------------------------------------
# WEB SOCKET SUPPORT
# -------------------------------------------------------------------

@app.websocket("/ws/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    """WebSocket endpoint для реального времени"""
    await websocket_manager.connect(websocket, client_id)
    try:
        while True:
            # Получаем сообщения от клиента
            data = await websocket.receive_json()
            
            # Обрабатываем команды
            if data.get("type") == "subscribe":
                batch_id = data.get("batch_id")
                if batch_id:
                    await websocket_manager.subscribe_to_batch(client_id, batch_id)
                    await websocket_manager.send_to_client(client_id, {
                        "type": "subscribed",
                        "batch_id": batch_id
                    })
            
            elif data.get("type") == "unsubscribe":
                batch_id = data.get("batch_id")
                if batch_id:
                    websocket_manager.unsubscribe_from_batch(client_id, batch_id)
                    await websocket_manager.send_to_client(client_id, {
                        "type": "unsubscribed",
                        "batch_id": batch_id
                    })
            
            elif data.get("type") == "ping":
                await websocket_manager.send_to_client(client_id, {
                    "type": "pong",
                    "timestamp": asyncio.get_event_loop().time()
                })
                
    except WebSocketDisconnect:
        websocket_manager.disconnect(client_id)
    except Exception as e:
        log.error(f"WebSocket error for client {client_id}: {e}")
        websocket_manager.disconnect(client_id)

# -------------------------------------------------------------------
# HEALTH CHECK
# -------------------------------------------------------------------

@app.get("/health")
async def health_check():
    """Проверка здоровья сервиса"""
    return {
        "status": "healthy",
        "service": "registry-control-api",
        "version": "1.0.0",
        "websocket_connections": len(websocket_manager.active_connections),
        "batch_subscriptions": len(websocket_manager.batch_subscriptions)
    }

@app.get("/health/redis")
async def redis_health():
    """Проверка Redis"""
    try:
        await redis_manager.ping()
        return {"redis": "connected"}
    except Exception as e:
        return {"redis": "disconnected", "error": str(e)}

# -------------------------------------------------------------------
# STARTUP/SHUTDOWN
# -------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)